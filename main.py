##  IMPORTS ##
#   GUI
import tkinter as tk
from tkinter import messagebox, filedialog
#   Selenium
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
#   Openpyxl
from openpyxl import Workbook, load_workbook

#   Misc.
import time


##  CODE ##
#   LOGIN FUNCTION
def login(username, password):
    service = Service()
    option = webdriver.ChromeOptions()
    #option.add_argument('--headless')
    driver = webdriver.Chrome(service=service, options=option)
    url = "https://www.twitter.com/login"
    driver.get(url)
    time.sleep(3)

    username_field = driver.find_element(By.TAG_NAME, "input")
    username_field.click()
    time.sleep(2)
    username_field.send_keys(username)
    page_button = driver.find_elements(By.XPATH, "//div[@role='button']")
    page_button[-2].click()
    
    time.sleep(2)
    pass_field = driver.find_element(By.XPATH, "//input[@type='password']")
    pass_field.send_keys(password)
    time.sleep(2) 
    page_button = driver.find_elements(By.XPATH, "//div[@role='button']")
    page_button[-1].click()

    return driver

#   TWEET FUNCTION
def tweet(driver, tweet_content):
    url = "https://twitter.com/compose/tweet"
    driver.get(url)
    time.sleep(2)
    tweet_field = driver.find_element(By.XPATH, "//div[@role='textbox']")
    tweet_field.send_keys(tweet_content)
    time.sleep(2)
    page_button = driver.find_element(By.XPATH, "//div[@role='button' and .//span[text()='Post']]")
    page_button.click()
    time.sleep(2)

def main(username, password, tweet_content):
    driver = login(username, password)
    time.sleep(2)
    tweet(driver, tweet_content)

def import_xlsx(filename):
    wb=load_workbook(filename)
    ws = wb.active
    max_row=ws.max_row
    for row in range(1,max_row+1):
        username=(ws['a' + str(row)].value)
        password=(ws['b' + str(row)].value)
        tweet_content=(ws['c' + str(row)].value)
        if username and password and tweet_content is not None:
            main(username, password, tweet_content)
        else:
            print("Error submitting content in line "+str(row))
            continue

def interface():
    
    def submit():
        username = username_entry.get()
        password = password_entry.get()
        tweet_content = tweet_text.get("1.0", "end-1c")
        if not username or not password or not tweet_content.strip():
            messagebox.showerror("Error", "All fields must be filled")
            return
        messagebox.showinfo("Information", "Tweet Submitted")
        main(username, password, tweet_content)

    def check_length(*args):
        tweet_content = tweet_text.get("1.0", "end-1c")
        count.set(f"Characters: {len(tweet_content)}/280")
        if len(tweet_content) > 280:
            tweet_text.delete("1.0", "end")
            tweet_text.insert("1.0", tweet_content[:280])

    def select_file():
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        import_xlsx(filename)

    
    root = tk.Tk()
    root.title("TweetBot")
    root.configure(bg='#1DA1F2')
    window_width = 800
    window_height = 600
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_top = int(screen_height / 2 - window_height / 2)
    position_right = int(screen_width / 2 - window_width / 2)
    root.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")


    header = tk.Label(root, text="TweetBot", font=("Arial", 30, "bold"), bg='#1DA1F2', fg='white')
    header.pack()
    username_label = tk.Label(root, text="Username", font=("Arial", 20, "bold"), bg='#1DA1F2', fg='white')
    username_label.pack()
    username_entry = tk.Entry(root, font=("Arial", 18))
    username_entry.pack()

    spacer = tk.Frame(root, height=20, bg='#1DA1F2')
    spacer.pack()

    password_label = tk.Label(root, text="Password", font=("Arial", 20, "bold"), bg='#1DA1F2', fg='white')
    password_label.pack()
    password_entry = tk.Entry(root, show="*", font=("Arial", 18))
    password_entry.pack()

    spacer = tk.Frame(root, height=20, bg='#1DA1F2')
    spacer.pack()

    tweet_label = tk.Label(root, text="Tweet", font=("Arial", 20, "bold"), bg='#1DA1F2', fg='white')
    tweet_label.pack()
    tweet_text = tk.Text(root, width=50, height=9, font=("Arial", 18))
    tweet_text.pack()

    count = tk.StringVar()
    count.set("Characters: 0/280")
    count_label = tk.Label(root, textvariable=count, font=("Arial", 15, "bold"), bg='#1DA1F2', fg='white')
    count_label.pack()

    tweet_text.bind("<KeyRelease>", check_length)

    button_frame = tk.Frame(root, bg='#1DA1F2')
    button_frame.pack()

    submit_button = tk.Button(button_frame, text="Submit", command=submit, font=("Arial", 18), fg='#1DA1F2', bg='white')
    submit_button.pack(side='left')

    import_button = tk.Button(button_frame, text="Import XLSX", command=select_file, font=("Arial", 18), fg='#1DA1F2', bg='white')
    import_button.pack(side='left')

    root.mainloop()

interface()